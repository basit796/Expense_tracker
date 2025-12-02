"""
Expense Tracker Backend with Firestore
This version uses Firebase Firestore instead of JSON files
Can be deployed to Render, Railway, or run locally
"""
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr
from typing import List, Optional
from datetime import datetime
import hashlib
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv
from google import genai
import os
import firebase_admin
from firebase_admin import credentials, firestore
from uuid import uuid4

load_dotenv()

# Initialize Firebase
cred = credentials.Certificate('serviceAccountKey.json')
firebase_admin.initialize_app(cred)
db = firestore.client()

app = FastAPI(title="Expense Tracker API with Firestore", version="3.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Update this to your Firebase Hosting URL later
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize Gemini AI
GOOGLE_API_KEY = os.getenv('GOOGLE_API_KEY')
if GOOGLE_API_KEY:
    client = genai.Client(api_key=GOOGLE_API_KEY)

# Models
class UserRegister(BaseModel):
    username: str
    email: EmailStr
    password: str
    fullName: str
    currency: Optional[str] = "PKR"

class UserLogin(BaseModel):
    username: str
    password: str

class TransactionCreate(BaseModel):
    username: str
    type: str
    category: str
    amount: float
    description: str
    date: str
    currency: Optional[str] = "PKR"

class BudgetCreate(BaseModel):
    username: str
    category: str
    amount: float
    month: str
    currency: Optional[str] = "PKR"

class GoalCreate(BaseModel):
    username: str
    name: str
    target_amount: float
    deadline: str
    currency: Optional[str] = "PKR"

class ContributeGoal(BaseModel):
    id: str
    amount: float

class SavingsOperation(BaseModel):
    username: str
    amount: float

class ChatMessage(BaseModel):
    username: str
    message: str

class UpdateName(BaseModel):
    username: str
    fullName: str

class UpdatePassword(BaseModel):
    username: str
    oldPassword: str
    newPassword: str

class UpdateCurrency(BaseModel):
    username: str
    currency: str

# Utility functions
def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

def generate_id() -> str:
    return str(uuid4())

# ========== USER ENDPOINTS ==========

@app.post("/api/register")
async def register(user: UserRegister):
    """Register a new user"""
    users_ref = db.collection('users')
    
    # Check if user exists
    existing_user = users_ref.document(user.username).get()
    if existing_user.exists:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Check email
    users = users_ref.where('email', '==', user.email).get()
    if len(list(users)) > 0:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Create user
    users_ref.document(user.username).set({
        'email': user.email,
        'password': hash_password(user.password),
        'fullName': user.fullName,
        'currency': user.currency,
        'savingsVault': 0,
        'createdAt': firestore.SERVER_TIMESTAMP
    })
    
    return {"message": "User registered successfully", "username": user.username}

@app.post("/api/login")
async def login(credentials: UserLogin):
    """User login"""
    user_ref = db.collection('users').document(credentials.username)
    user_doc = user_ref.get()
    
    if not user_doc.exists:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    user_data = user_doc.to_dict()
    if user_data['password'] != hash_password(credentials.password):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    return {
        "message": "Login successful",
        "username": credentials.username,
        "email": user_data['email'],
        "fullName": user_data['fullName'],
        "currency": user_data.get('currency', 'PKR')
    }

@app.get("/api/profile/{username}")
async def get_profile(username: str):
    """Get user profile"""
    user_ref = db.collection('users').document(username)
    user_doc = user_ref.get()
    
    if not user_doc.exists:
        raise HTTPException(status_code=404, detail="User not found")
    
    user_data = user_doc.to_dict()
    return {
        "username": username,
        "email": user_data['email'],
        "fullName": user_data['fullName'],
        "currency": user_data.get('currency', 'PKR'),
        "savingsVault": user_data.get('savingsVault', 0)
    }

@app.put("/api/profile/name")
async def update_name(data: UpdateName):
    """Update user full name"""
    user_ref = db.collection('users').document(data.username)
    if not user_ref.get().exists:
        raise HTTPException(status_code=404, detail="User not found")
    
    user_ref.update({'fullName': data.fullName})
    return {"message": "Name updated successfully"}

@app.put("/api/profile/password")
async def update_password(data: UpdatePassword):
    """Update user password"""
    user_ref = db.collection('users').document(data.username)
    user_doc = user_ref.get()
    
    if not user_doc.exists:
        raise HTTPException(status_code=404, detail="User not found")
    
    user_data = user_doc.to_dict()
    if user_data['password'] != hash_password(data.oldPassword):
        raise HTTPException(status_code=400, detail="Old password is incorrect")
    
    user_ref.update({'password': hash_password(data.newPassword)})
    return {"message": "Password updated successfully"}

@app.put("/api/profile/currency")
async def update_currency(data: UpdateCurrency):
    """Update user currency preference"""
    user_ref = db.collection('users').document(data.username)
    if not user_ref.get().exists:
        raise HTTPException(status_code=404, detail="User not found")
    
    user_ref.update({'currency': data.currency})
    return {"message": "Currency updated successfully"}

# ========== TRANSACTION ENDPOINTS ==========

@app.post("/api/transactions")
async def add_transaction(txn: TransactionCreate):
    """Add a new transaction"""
    txn_id = generate_id()
    db.collection('transactions').document(txn_id).set({
        'id': txn_id,
        'username': txn.username,
        'type': txn.type,
        'category': txn.category,
        'amount': txn.amount,
        'description': txn.description,
        'date': txn.date,
        'currency': txn.currency,
        'created_at': datetime.now().isoformat()
    })
    return {"message": "Transaction added successfully", "id": txn_id}

@app.get("/api/transactions/{username}")
async def get_transactions(username: str):
    """Get all transactions for a user"""
    transactions = db.collection('transactions').where('username', '==', username).stream()
    return [txn.to_dict() for txn in transactions]

@app.delete("/api/transactions/{txn_id}")
async def delete_transaction(txn_id: str):
    """Delete a transaction"""
    db.collection('transactions').document(txn_id).delete()
    return {"message": "Transaction deleted successfully"}

# ========== REPORT ENDPOINTS ==========

@app.get("/api/report/{username}")
async def get_monthly_report(username: str, month: Optional[str] = None):
    """Get monthly financial report"""
    transactions = db.collection('transactions').where('username', '==', username).stream()
    txn_list = [txn.to_dict() for txn in transactions]
    
    # Filter by month if provided
    if month:
        txn_list = [t for t in txn_list if t['date'].startswith(month)]
    
    total_income = sum(t['amount'] for t in txn_list if t['type'] == 'income')
    total_expense = sum(t['amount'] for t in txn_list if t['type'] == 'expense')
    
    category_breakdown = {}
    for txn in txn_list:
        if txn['type'] == 'expense':
            category_breakdown[txn['category']] = category_breakdown.get(txn['category'], 0) + txn['amount']
    
    # Get user savings
    user_ref = db.collection('users').document(username)
    user_data = user_ref.get().to_dict()
    
    return {
        "total_income": total_income,
        "total_expense": total_expense,
        "balance": total_income - total_expense,
        "category_breakdown": category_breakdown,
        "transaction_count": len(txn_list),
        "savingsVault": user_data.get('savingsVault', 0)
    }

# ========== SAVINGS VAULT ENDPOINTS ==========

@app.post("/api/savings/add")
async def add_to_savings(data: SavingsOperation):
    """Add money to savings vault"""
    user_ref = db.collection('users').document(data.username)
    user_doc = user_ref.get()
    
    if not user_doc.exists:
        raise HTTPException(status_code=404, detail="User not found")
    
    current_savings = user_doc.to_dict().get('savingsVault', 0)
    user_ref.update({'savingsVault': current_savings + data.amount})
    
    return {"message": "Added to savings", "newBalance": current_savings + data.amount}

@app.post("/api/savings/withdraw")
async def withdraw_from_savings(data: SavingsOperation):
    """Withdraw from savings vault"""
    user_ref = db.collection('users').document(data.username)
    user_doc = user_ref.get()
    
    if not user_doc.exists:
        raise HTTPException(status_code=404, detail="User not found")
    
    current_savings = user_doc.to_dict().get('savingsVault', 0)
    if data.amount > current_savings:
        raise HTTPException(status_code=400, detail="Insufficient savings")
    
    user_ref.update({'savingsVault': current_savings - data.amount})
    return {"message": "Withdrawn from savings", "newBalance": current_savings - data.amount}

# ========== BUDGET ENDPOINTS ==========

@app.post("/api/budgets/set")
async def set_budget(budget: BudgetCreate):
    """Set a budget for a category"""
    budget_id = f"{budget.username}_{budget.category}_{budget.month}"
    db.collection('budgets').document(budget_id).set({
        'id': budget_id,
        'username': budget.username,
        'category': budget.category,
        'amount': budget.amount,
        'month': budget.month,
        'currency': budget.currency,
        'createdAt': firestore.SERVER_TIMESTAMP
    })
    return {"message": "Budget set successfully"}

@app.get("/api/budgets/status/{username}")
async def get_budget_status(username: str, month: Optional[str] = None):
    """Get budget status with spending"""
    if not month:
        month = datetime.now().strftime("%Y-%m")
    
    budgets = db.collection('budgets').where('username', '==', username).where('month', '==', month).stream()
    budget_list = [b.to_dict() for b in budgets]
    
    transactions = db.collection('transactions').where('username', '==', username).stream()
    txn_list = [t.to_dict() for t in transactions if t.to_dict()['date'].startswith(month) and t.to_dict()['type'] == 'expense']
    
    budget_status = []
    for budget in budget_list:
        spent = sum(t['amount'] for t in txn_list if t['category'] == budget['category'])
        remaining = budget['amount'] - spent
        percentage = (spent / budget['amount'] * 100) if budget['amount'] > 0 else 0
        
        status = 'good'
        if percentage >= 100:
            status = 'exceeded'
        elif percentage >= 80:
            status = 'warning'
        
        budget_status.append({
            'category': budget['category'],
            'budget': budget['amount'],
            'spent': spent,
            'remaining': remaining,
            'percentage': percentage,
            'status': status,
            'currency': budget['currency']
        })
    
    return {"budget_status": budget_status}

@app.delete("/api/budgets/{budget_id}")
async def delete_budget(budget_id: str):
    """Delete a budget"""
    db.collection('budgets').document(budget_id).delete()
    return {"message": "Budget deleted successfully"}

# ========== GOALS ENDPOINTS ==========

@app.post("/api/goals/create")
async def create_goal(goal: GoalCreate):
    """Create a financial goal"""
    goal_id = generate_id()
    
    # Calculate days remaining
    deadline_date = datetime.strptime(goal.deadline, "%Y-%m-%d")
    days_remaining = (deadline_date - datetime.now()).days
    
    db.collection('goals').document(goal_id).set({
        'id': goal_id,
        'username': goal.username,
        'name': goal.name,
        'target_amount': goal.target_amount,
        'current_amount': 0,
        'deadline': goal.deadline,
        'currency': goal.currency,
        'createdAt': firestore.SERVER_TIMESTAMP
    })
    
    return {"message": "Goal created successfully", "id": goal_id}

@app.get("/api/goals/{username}")
async def get_goals(username: str):
    """Get all financial goals for a user"""
    goals = db.collection('goals').where('username', '==', username).stream()
    goal_list = []
    
    for goal_doc in goals:
        goal = goal_doc.to_dict()
        
        # Calculate progress
        progress = (goal['current_amount'] / goal['target_amount'] * 100) if goal['target_amount'] > 0 else 0
        remaining = goal['target_amount'] - goal['current_amount']
        
        # Calculate days remaining
        deadline_date = datetime.strptime(goal['deadline'], "%Y-%m-%d")
        days_remaining = max(0, (deadline_date - datetime.now()).days)
        
        # Calculate daily savings required
        daily_required = remaining / days_remaining if days_remaining > 0 else 0
        
        goal_list.append({
            'id': goal['id'],
            'name': goal['name'],
            'target_amount': goal['target_amount'],
            'current_amount': goal['current_amount'],
            'progress_percentage': progress,
            'remaining': remaining,
            'deadline': goal['deadline'],
            'days_remaining': days_remaining,
            'daily_savings_required': daily_required,
            'category': 'savings',
            'currency': goal['currency']
        })
    
    return goal_list

@app.post("/api/goals/contribute")
async def contribute_to_goal(data: ContributeGoal):
    """Contribute money to a goal"""
    goal_ref = db.collection('goals').document(data.id)
    goal_doc = goal_ref.get()
    
    if not goal_doc.exists:
        raise HTTPException(status_code=404, detail="Goal not found")
    
    goal_data = goal_doc.to_dict()
    new_amount = goal_data['current_amount'] + data.amount
    
    goal_ref.update({'current_amount': new_amount})
    return {"message": "Contribution added successfully"}

@app.delete("/api/goals/{goal_id}")
async def delete_goal(goal_id: str):
    """Delete a financial goal"""
    db.collection('goals').document(goal_id).delete()
    return {"message": "Goal deleted successfully"}

# ========== CURRENCY ENDPOINTS ==========

@app.get("/api/currency/rates")
async def get_currency_rates():
    """Get currency exchange rates"""
    rates_ref = db.collection('currency_rates').document('rates')
    rates_doc = rates_ref.get()
    
    if not rates_doc.exists:
        # Default rates
        default_rates = {
            "USD": 1.0,
            "EUR": 0.85,
            "GBP": 0.73,
            "PKR": 278.50,
            "SAR": 3.75,
            "AED": 3.67
        }
        rates_ref.set(default_rates)
        return {"rates": default_rates}
    
    return {"rates": rates_doc.to_dict()}

# ========== EXPORT ENDPOINTS ==========

@app.get("/api/export/{username}")
async def export_to_excel(username: str, month: Optional[str] = None):
    """Export transactions to Excel"""
    transactions = db.collection('transactions').where('username', '==', username).stream()
    txn_list = [txn.to_dict() for txn in transactions]
    
    if month:
        txn_list = [t for t in txn_list if t['date'].startswith(month)]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    
    headers = ["Date", "Type", "Category", "Amount", "Currency", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row, txn in enumerate(txn_list, 2):
        ws.cell(row=row, column=1, value=txn['date'])
        ws.cell(row=row, column=2, value=txn['type'])
        ws.cell(row=row, column=3, value=txn['category'])
        ws.cell(row=row, column=4, value=txn['amount'])
        ws.cell(row=row, column=5, value=txn.get('currency', 'PKR'))
        ws.cell(row=row, column=6, value=txn['description'])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"transactions_{username}_{month or 'all'}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ========== AI CHAT ENDPOINTS ==========

@app.post("/api/chat")
async def chat_with_ai(message: ChatMessage):
    """Chat with AI assistant"""
    if not GOOGLE_API_KEY:
        raise HTTPException(status_code=500, detail="AI service not configured")
    
    # Get user's financial context
    transactions = db.collection('transactions').where('username', '==', message.username).stream()
    txn_list = [txn.to_dict() for txn in transactions]
    
    total_income = sum(t['amount'] for t in txn_list if t['type'] == 'income')
    total_expense = sum(t['amount'] for t in txn_list if t['type'] == 'expense')
    
    context = f"""
    User's Financial Summary:
    - Total Income: ${total_income}
    - Total Expenses: ${total_expense}
    - Balance: ${total_income - total_expense}
    - Number of Transactions: {len(txn_list)}
    
    User Question: {message.message}
    
    Provide helpful financial advice based on this data.
    """
    
    response = client.models.generate_content(
        model='gemini-2.0-flash-exp',
        contents=context
    )
    
    return {
        "response": response.text,
        "context": {
            "income": total_income,
            "expense": total_expense,
            "balance": total_income - total_expense
        }
    }

@app.get("/")
async def root():
    return {
        "message": "Expense Tracker API with Firestore",
        "version": "3.0.0",
        "status": "running",
        "database": "Firestore"
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
