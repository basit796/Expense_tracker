from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr
from typing import List, Optional
from datetime import datetime
import hashlib
from pathlib import Path
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import httpx
import asyncio
from dotenv import load_dotenv
from google import genai
import os
import json

load_dotenv()

app = FastAPI(title="Expense Tracker API", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

JAVA_BACKEND_URL = "http://localhost:9000"

DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)
USERS_FILE = DATA_DIR / "users.json"
TRANSACTIONS_FILE = DATA_DIR / "transactions.json"

if not USERS_FILE.exists():
    USERS_FILE.write_text("[]")
if not TRANSACTIONS_FILE.exists():
    TRANSACTIONS_FILE.write_text("[]")

# Models
class UserRegister(BaseModel):
    username: str
    email: EmailStr
    password: str
    fullName: str
    currency: Optional[str] = "PKR"

class UserLogin(BaseModel):
    username: str
    password: str

class TransactionCreate(BaseModel):
    username: str
    type: str
    category: str
    amount: float
    description: str
    date: str
    currency: Optional[str] = "PKR"

class AnalysisRequest(BaseModel):
    username: str
    query: str

class ChatMessage(BaseModel):
    username: str
    message: str

# Utility functions
def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

def load_users() -> List[dict]:
    return json.loads(USERS_FILE.read_text())

def save_users(users: List[dict]):
    USERS_FILE.write_text(json.dumps(users, indent=2))

def load_transactions() -> List[dict]:
    return json.loads(TRANSACTIONS_FILE.read_text())

def save_transactions(transactions: List[dict]):
    TRANSACTIONS_FILE.write_text(json.dumps(transactions, indent=2))

def generate_id() -> str:
    from uuid import uuid4
    return str(uuid4())

def load_local_transactions():
    """Load transactions from Java backend JSON file"""
    try:
        # Use Java backend data directory
        data_path = Path(__file__).parent.parent / "java-backend" / "data" / "transactions.json"
        if data_path.exists():
            content = data_path.read_text()
            transactions = json.loads(content)
            return transactions
        return []
    except Exception as e:
        print(f"Error reading transactions from Java backend: {e}")
        return []

def load_users_from_java():
    """Load users from Java backend JSON file"""
    try:
        data_path = Path(__file__).parent.parent / "java-backend" / "data" / "users.json"
        if data_path.exists():
            content = data_path.read_text()
            users = json.loads(content)
            return users
        return []
    except Exception as e:
        print(f"Error reading users from Java backend: {e}")
        return []

# Root endpoint
@app.get("/")
def read_root():
    return {"message": "Expense Tracker API", "version": "2.0.0", "status": "running"}

# All your existing endpoints...
@app.post("/api/register")
async def register(user: UserRegister):
    async with httpx.AsyncClient() as client:
        try:
            response = await client.post(
                f"{JAVA_BACKEND_URL}/api/java/register",
                json=user.dict(),
                timeout=10.0
            )
            if response.status_code == 200:
                return response.json()
            else:
                raise HTTPException(status_code=response.status_code, detail=response.json().get("error", "Registration failed"))
        except httpx.RequestError:
            raise HTTPException(status_code=503, detail="Java backend unavailable")

@app.post("/api/login")
async def login(user: UserLogin):
    async with httpx.AsyncClient() as client:
        try:
            response = await client.post(
                f"{JAVA_BACKEND_URL}/api/java/login",
                json=user.dict(),
                timeout=10.0
            )
            if response.status_code == 200:
                return response.json()
            else:
                raise HTTPException(status_code=response.status_code, detail=response.json().get("error", "Invalid credentials"))
        except httpx.RequestError:
            raise HTTPException(status_code=503, detail="Java backend unavailable")

@app.get("/api/profile/{username}")
async def get_profile(username: str):
    async with httpx.AsyncClient() as client:
        try:
            response = await client.get(
                f"{JAVA_BACKEND_URL}/api/java/profile",
                params={"username": username},
                timeout=10.0
            )
            if response.status_code == 200:
                return response.json()
            else:
                raise HTTPException(status_code=response.status_code, detail=response.json().get("error", "User not found"))
        except httpx.RequestError:
            raise HTTPException(status_code=503, detail="Java backend unavailable")

@app.put("/api/profile/name")
async def update_name(data: dict):
    async with httpx.AsyncClient() as client:
        try:
            response = await client.put(
                f"{JAVA_BACKEND_URL}/api/java/profile/name",
                json=data,
                timeout=10.0
            )
            if response.status_code == 200:
                return response.json()
            else:
                raise HTTPException(status_code=response.status_code, detail=response.json().get("error", "Update failed"))
        except httpx.RequestError:
            raise HTTPException(status_code=503, detail="Java backend unavailable")

@app.put("/api/profile/password")
async def update_password(data: dict):
    async with httpx.AsyncClient() as client:
        try:
            response = await client.put(
                f"{JAVA_BACKEND_URL}/api/java/profile/password",
                json=data,
                timeout=10.0
            )
            if response.status_code == 200:
                return response.json()
            else:
                raise HTTPException(status_code=response.status_code, detail=response.json().get("error", "Password update failed"))
        except httpx.RequestError:
            raise HTTPException(status_code=503, detail="Java backend unavailable")

@app.put("/api/profile/currency")
async def update_currency(data: dict):
    async with httpx.AsyncClient() as client:
        try:
            response = await client.put(
                f"{JAVA_BACKEND_URL}/api/java/profile/currency",
                json=data,
                timeout=10.0
            )
            if response.status_code == 200:
                return response.json()
            else:
                raise HTTPException(status_code=response.status_code, detail=response.json().get("error", "Currency update failed"))
        except httpx.RequestError:
            raise HTTPException(status_code=503, detail="Java backend unavailable")

@app.post("/api/transactions")
async def add_transaction(transaction: TransactionCreate):
    async with httpx.AsyncClient() as client:
        try:
            response = await client.post(
                f"{JAVA_BACKEND_URL}/api/java/transactions/add",
                json=transaction.dict(),
                timeout=10.0
            )
            if response.status_code == 200:
                return response.json()
            else:
                raise HTTPException(status_code=response.status_code, detail=response.json().get("error", "Transaction failed"))
        except httpx.RequestError:
            raise HTTPException(status_code=503, detail="Java backend unavailable")

@app.get("/api/transactions/{username}")
async def get_transactions(username: str):
    async with httpx.AsyncClient() as client:
        try:
            response = await client.get(
                f"{JAVA_BACKEND_URL}/api/java/transactions/get",
                params={"username": username},
                timeout=10.0
            )
            if response.status_code == 200:
                return response.json()
            else:
                raise HTTPException(status_code=response.status_code, detail=response.json().get("error", "Failed to fetch transactions"))
        except httpx.RequestError:
            raise HTTPException(status_code=503, detail="Java backend unavailable")

@app.get("/api/report/{username}")
async def get_monthly_report(username: str, month: Optional[str] = None):
    async with httpx.AsyncClient() as client:
        try:
            params = {"username": username}
            if month:
                params["month"] = month
            response = await client.get(
                f"{JAVA_BACKEND_URL}/api/java/report",
                params=params,
                timeout=10.0
            )
            if response.status_code == 200:
                return response.json()
            else:
                raise HTTPException(status_code=response.status_code, detail=response.json().get("error", "Report generation failed"))
        except httpx.RequestError:
            raise HTTPException(status_code=503, detail="Java backend unavailable")

@app.get("/api/currency/rates")
async def get_currency_rates():
    async with httpx.AsyncClient() as client:
        try:
            response = await client.get(
                f"{JAVA_BACKEND_URL}/api/java/currency/rates",
                timeout=10.0
            )
            if response.status_code == 200:
                return response.json()
            else:
                raise HTTPException(status_code=response.status_code, detail="Failed to fetch rates")
        except httpx.RequestError:
            raise HTTPException(status_code=503, detail="Java backend unavailable")

@app.get("/api/export/{username}")
async def export_excel(username: str, month: Optional[str] = None):
    async with httpx.AsyncClient() as client:
        try:
            params = {"username": username}
            if month:
                params["month"] = month
            response = await client.get(
                f"{JAVA_BACKEND_URL}/api/java/transactions/get",
                params=params,
                timeout=10.0
            )
            
            if response.status_code != 200:
                raise HTTPException(status_code=404, detail="User not found")
            
            user_transactions = response.json()
            
            if month:
                user_transactions = [t for t in user_transactions if t["date"].startswith(month)]
        except httpx.RequestError:
            raise HTTPException(status_code=503, detail="Java backend unavailable")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = ["Date", "Type", "Category", "Amount", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for row, t in enumerate(user_transactions, 2):
        ws.cell(row=row, column=1, value=t["date"])
        ws.cell(row=row, column=2, value=t["type"].capitalize())
        ws.cell(row=row, column=3, value=t["category"])
        ws.cell(row=row, column=4, value=t["amount"])
        ws.cell(row=row, column=5, value=t["description"])
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 30
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"expense_report_{username}_{month or 'all'}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.delete("/api/transactions/{transaction_id}")
async def delete_transaction(transaction_id: str):
    async with httpx.AsyncClient() as client:
        try:
            response = await client.delete(
                f"{JAVA_BACKEND_URL}/api/java/transactions/delete",
                params={"id": transaction_id},
                timeout=10.0
            )
            if response.status_code == 200:
                return response.json()
            else:
                raise HTTPException(status_code=response.status_code, detail=response.json().get("error", "Delete failed"))
        except httpx.RequestError:
            raise HTTPException(status_code=503, detail="Java backend unavailable")

@app.post("/api/savings/add")
async def add_to_savings(data: dict):
    async with httpx.AsyncClient() as client:
        try:
            response = await client.post(
                f"{JAVA_BACKEND_URL}/api/java/savings/add",
                json=data,
                timeout=10.0
            )
            if response.status_code == 200:
                return response.json()
            else:
                raise HTTPException(status_code=response.status_code, detail=response.json().get("error", "Failed to add to savings"))
        except httpx.RequestError:
            raise HTTPException(status_code=503, detail="Java backend unavailable")

@app.post("/api/savings/withdraw")
async def withdraw_from_savings(data: dict):
    async with httpx.AsyncClient() as client:
        try:
            response = await client.post(
                f"{JAVA_BACKEND_URL}/api/java/savings/withdraw",
                json=data,
                timeout=10.0
            )
            if response.status_code == 200:
                return response.json()
            else:
                raise HTTPException(status_code=response.status_code, detail=response.json().get("error", "Failed to withdraw from savings"))
        except httpx.RequestError:
            raise HTTPException(status_code=503, detail="Java backend unavailable")


# --- Budget Management Endpoints ---

@app.post("/api/budgets/set")
async def set_budget(data: dict):
    async with httpx.AsyncClient() as client:
        try:
            response = await client.post(
                f"{JAVA_BACKEND_URL}/api/java/budgets/set",
                json=data,
                timeout=10.0
            )
            if response.status_code == 200:
                return response.json()
            else:
                raise HTTPException(status_code=response.status_code, detail=response.json().get("error", "Failed to set budget"))
        except httpx.RequestError:
            raise HTTPException(status_code=503, detail="Java backend unavailable")

@app.get("/api/budgets/{username}")
async def get_budgets(username: str, month: Optional[str] = None):
    async with httpx.AsyncClient() as client:
        try:
            params = {"username": username}
            if month:
                params["month"] = month
            response = await client.get(
                f"{JAVA_BACKEND_URL}/api/java/budgets/get",
                params=params,
                timeout=10.0
            )
            if response.status_code == 200:
                return response.json()
            else:
                raise HTTPException(status_code=response.status_code, detail="Failed to fetch budgets")
        except httpx.RequestError:
            raise HTTPException(status_cxode=503, detail="Java backend unavailable")

@app.get("/api/budgets/status/{username}")
async def get_budget_status(username: str, month: Optional[str] = None):
    async with httpx.AsyncClient() as client:
        try:
            params = {"username": username}
            if month:
                params["month"] = month
            response = await client.get(
                f"{JAVA_BACKEND_URL}/api/java/budgets/status",
                params=params,
                timeout=10.0
            )
            if response.status_code == 200:
                return response.json()
            else:
                raise HTTPException(status_code=response.status_code, detail="Failed to fetch budget status")
        except httpx.RequestError:
            raise HTTPException(status_code=503, detail="Java backend unavailable")

@app.delete("/api/budgets/{budget_id}")
async def delete_budget(budget_id: str):
    async with httpx.AsyncClient() as client:
        try:
            response = await client.delete(
                f"{JAVA_BACKEND_URL}/api/java/budgets/delete",
                params={"id": budget_id},
                timeout=10.0
            )
            if response.status_code == 200:
                return response.json()
            else:
                raise HTTPException(status_code=response.status_code, detail="Failed to delete budget")
        except httpx.RequestError:
            raise HTTPException(status_code=503, detail="Java backend unavailable")


# --- Financial Goals Endpoints ---

@app.post("/api/goals/create")
async def create_goal(data: dict):
    async with httpx.AsyncClient() as client:
        try:
            response = await client.post(
                f"{JAVA_BACKEND_URL}/api/java/goals/create",
                json=data,
                timeout=10.0
            )
            if response.status_code == 200:
                return response.json()
            else:
                raise HTTPException(status_code=response.status_code, detail=response.json().get("error", "Failed to create goal"))
        except httpx.RequestError:
            raise HTTPException(status_code=503, detail="Java backend unavailable")

@app.get("/api/goals/{username}")
async def get_goals(username: str):
    async with httpx.AsyncClient() as client:
        try:
            response = await client.get(
                f"{JAVA_BACKEND_URL}/api/java/goals/get",
                params={"username": username},
                timeout=10.0
            )
            if response.status_code == 200:
                return response.json()
            else:
                raise HTTPException(status_code=response.status_code, detail="Failed to fetch goals")
        except httpx.RequestError:
            raise HTTPException(status_code=503, detail="Java backend unavailable")

@app.post("/api/goals/contribute")
async def contribute_to_goal(data: dict):
    async with httpx.AsyncClient() as client:
        try:
            response = await client.post(
                f"{JAVA_BACKEND_URL}/api/java/goals/contribute",
                json=data,
                timeout=10.0
            )
            if response.status_code == 200:
                return response.json()
            else:
                raise HTTPException(status_code=response.status_code, detail=response.json().get("error", "Failed to contribute to goal"))
        except httpx.RequestError:
            raise HTTPException(status_code=503, detail="Java backend unavailable")

@app.delete("/api/goals/{goal_id}")
async def delete_goal(goal_id: str):
    async with httpx.AsyncClient() as client:
        try:
            response = await client.delete(
                f"{JAVA_BACKEND_URL}/api/java/goals/delete",
                params={"id": goal_id},
                timeout=10.0
            )
            if response.status_code == 200:
                return response.json()
            else:
                raise HTTPException(status_code=response.status_code, detail="Failed to delete goal")
        except httpx.RequestError:
            raise HTTPException(status_code=503, detail="Java backend unavailable")


# --- AI Analysis Endpoint (Direct Gemini Integration) ---

@app.post("/api/analyze")
async def analyze_expenses_endpoint(request: AnalysisRequest):
    """
    Direct endpoint to test AI analysis without CopilotKit
    Test with: POST http://localhost:8000/api/analyze
    Body: {"username": "your_username", "query": "What are my spending patterns?"}
    """
    api_key = os.getenv("GOOGLE_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="GOOGLE_API_KEY not configured")
    
    try:
        # Load user's transactions
        all_transactions = load_local_transactions()
        user_transactions = [t for t in all_transactions if t.get("username") == request.username]
        
        if not user_transactions:
            return {
                "analysis": f"No transactions found for user '{request.username}'. Please add some transactions first.",
                "transaction_count": 0
            }
        
        # Create context for Gemini
        transactions_summary = json.dumps(user_transactions, indent=2)
        
        system_instruction = f"""You are a helpful financial assistant analyzing expense data.

User's Transaction History ({len(user_transactions)} transactions):
{transactions_summary}

User's Question: {request.query}

Provide a helpful analysis based on their data. Include:
- Direct answer to their question
- Spending patterns if relevant
- Category breakdowns if relevant
- Income vs expenses if relevant
- Actionable recommendations

Be concise, helpful, and data-driven."""

        # Call Gemini API
        client = genai.Client(api_key=api_key)
        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=system_instruction,
        )
        
        return {
            "analysis": response.text,
            "transaction_count": len(user_transactions),
            "username": request.username
        }
        
    except Exception as e:
        print(f"Error in analysis: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Analysis failed: {str(e)}")


@app.post("/api/chat")
async def chat_endpoint(request: ChatMessage):
    """
    Chat endpoint for CopilotKit integration
    Analyzes user's financial data and responds to questions
    """
    api_key = os.getenv("GOOGLE_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="GOOGLE_API_KEY not configured")
    
    try:
        # Load user's transactions from Java backend
        all_transactions = load_local_transactions()
        user_transactions = [t for t in all_transactions if t.get("username") == request.username]
        
        # Load user profile from Java backend
        users = load_users_from_java()
        user_profile = next((u for u in users if u.get("username") == request.username), None)
        
        if not user_profile:
            return {
                "response": f"User profile not found. Please make sure you're logged in.",
                "context": {}
            }
        
        # Calculate financial summary
        total_income = sum(t.get("amount", 0) for t in user_transactions if t.get("type") == "income")
        total_expense = sum(t.get("amount", 0) for t in user_transactions if t.get("type") == "expense")
        balance = total_income - total_expense
        savings = user_profile.get("savingsVault", 0)
        currency = user_profile.get("currency", "PKR")
        
        # Category breakdown for expenses
        category_spending = {}
        for t in user_transactions:
            if t.get("type") == "expense":
                category = t.get("category", "Other")
                category_spending[category] = category_spending.get(category, 0) + t.get("amount", 0)
        
        # Find top spending category
        top_category = max(category_spending.items(), key=lambda x: x[1]) if category_spending else ("None", 0)
        
        # Recent transactions (last 5)
        recent = sorted(user_transactions, key=lambda x: x.get("created_at", ""), reverse=True)[:5]
        
        # Prepare context
        context = {
            "total_transactions": len(user_transactions),
            "total_income": total_income,
            "total_expense": total_expense,
            "balance": balance,
            "savings_vault": savings,
            "currency": currency,
            "category_breakdown": category_spending,
            "top_spending_category": top_category[0],
            "top_spending_amount": top_category[1],
            "recent_transactions": recent
        }
        
        # Build comprehensive financial summary for AI
        category_list = "\n".join([f"  • {cat}: {currency} {amt:,.2f}" for cat, amt in category_spending.items()])
        recent_list = "\n".join([f"  • {t.get('date', 'N/A')}: {t.get('category', 'N/A')} - {currency} {t.get('amount', 0):,.2f} ({t.get('type', 'expense')})" for t in recent[:5]])
        
        system_instruction = f"""You are a friendly and professional financial assistant for an expense tracking app.

**User's Complete Financial Summary:**

📊 **Overall Finances:**
- Total Income: {currency} {total_income:,.2f}
- Total Expenses: {currency} {total_expense:,.2f}
- Current Balance: {currency} {balance:,.2f}
- Savings Vault: {currency} {savings:,.2f}
- Total Transactions: {len(user_transactions)}
- Currency: {currency}

💰 **Spending by Category:**
{category_list if category_list else "  • No expenses recorded yet"}

🏆 **Top Spending Category:** {top_category[0]} ({currency} {top_category[1]:,.2f})

📝 **Recent Transactions (Last 5):**
{recent_list if recent_list else "  • No recent transactions"}

💡 **Financial Health:**
- Savings Rate: {((savings / total_income * 100) if total_income > 0 else 0):.1f}%
- Expense Ratio: {((total_expense / total_income * 100) if total_income > 0 else 0):.1f}%

**User's Question:** {request.message}

**Guidelines for your response:**
1. Be warm, conversational, and encouraging
2. Provide specific insights based on their ACTUAL data shown above
3. If they ask about spending, reference the category breakdown
4. If they ask for advice, give actionable financial tips based on their patterns
5. If they greet you, greet them back warmly and offer to help with their finances
6. Keep responses concise (2-4 paragraphs max)
7. Use their currency ({currency}) in all amounts
8. Be positive while being honest about areas for improvement
9. Analyze the data yourself - don't ask them for information you already have
10. Provide percentages and comparisons where relevant

**Important:** You have complete access to their financial data. Analyze it yourself and provide insights without asking them for more information.

Respond naturally and helpfully to their message."""

        # Call Gemini API
        client = genai.Client(api_key=api_key)
        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=system_instruction,
        )
        
        return {
            "response": response.text,
            "context": context
        }
        
    except Exception as e:
        print(f"Error in chat: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Chat failed: {str(e)}")


# --- CopilotKit Integration (Optional) ---
# try:
#     from copilotkit.integrations.api import add_fastapi_endpoint
#     from copilotkit import CopilotKitSDK, Action

#     async def copilot_analyze_expenses(username: str, query: str):
#         """CopilotKit action handler"""
#         api_key = os.getenv("GOOGLE_API_KEY")
#         if not api_key:
#             return {"error": "API key not configured"}
        
#         client = genai.Client(api_key=api_key)
#         all_transactions = load_local_transactions()
#         user_transactions = [t for t in all_transactions if t.get("username") == username]
        
#         transactions_summary = json.dumps(user_transactions, indent=2)
        
#         system_instruction = f"""You are a helpful financial assistant.

# Transaction History:
# {transactions_summary}

# Question: {query}

# Provide insights about spending patterns, categories, income vs expenses, and recommendations."""

#         try:
#             response = client.models.generate_content(
#                 model="gemini-2.5-pro",
#                 contents=system_instruction,
#             )
#             return {"analysis": response.text}
#         except Exception as e:
#             return {"error": f"Analysis failed: {str(e)}"}

#     expense_analysis_action = Action(
#         name="analyze_expenses",
#         description="Analyze user's expenses and provide financial insights",
#         handler=copilot_analyze_expenses,
#         parameters=[
#             {"name": "username", "type": "string", "description": "Username of the user"},
#             {"name": "query", "type": "string", "description": "The financial question"}
#         ]
#     )

#     sdk = CopilotKitSDK(actions=[expense_analysis_action])
#     add_fastapi_endpoint(app, sdk, "/copilotkit")
#     print("✅ CopilotKit integration enabled at /copilotkit")
    
# except ImportError:
#     print("⚠️  CopilotKit not installed. Only direct API endpoints available.")


if __name__ == "__main__":
    import uvicorn
    print("🚀 Starting Expense Tracker FastAPI...")
    print("📍 API: http://localhost:8000")
    print("📚 Docs: http://localhost:8000/docs")
    print("🔗 Connecting to Java Backend on port 9000")
    print("🤖 AI Analysis: POST /api/analyze")
    uvicorn.run(app, host="0.0.0.0", port=8000)