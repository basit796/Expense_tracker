"""
Expense Tracker Backend with Firestore - UUID-Based Schema
Uses userId (UUID) as primary identifier for better database design
"""
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr
from typing import List, Optional
from datetime import datetime
import hashlib
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv
from google import genai
import os
import firebase_admin
from firebase_admin import credentials, firestore
from uuid import uuid4

load_dotenv()

# Initialize Firebase
cred = credentials.Certificate('serviceAccountKey.json')
firebase_admin.initialize_app(cred)
db = firestore.client()

app = FastAPI(title="Expense Tracker API with Firestore (UUID)", version="3.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize Gemini AI
GOOGLE_API_KEY = os.getenv('GOOGLE_API_KEY')
if GOOGLE_API_KEY:
    client = genai.Client(api_key=GOOGLE_API_KEY)

# Models
class UserRegister(BaseModel):
    username: str
    email: EmailStr
    password: str
    fullName: str
    currency: Optional[str] = "PKR"

class UserLogin(BaseModel):
    username: str
    password: str

class TransactionCreate(BaseModel):
    username: str  # Frontend still sends username
    type: str
    category: str
    amount: float
    description: str
    date: str
    currency: Optional[str] = "PKR"

class BudgetCreate(BaseModel):
    username: str
    category: str
    amount: float
    month: str
    currency: Optional[str] = "PKR"

class GoalCreate(BaseModel):
    username: str
    name: str
    target_amount: float
    deadline: str
    currency: Optional[str] = "PKR"

class ContributeGoal(BaseModel):
    id: str
    amount: float

class SavingsOperation(BaseModel):
    username: str
    amount: float

class ChatMessage(BaseModel):
    username: str
    message: str

class UpdateName(BaseModel):
    username: str
    fullName: str

class UpdatePassword(BaseModel):
    username: str
    oldPassword: str
    newPassword: str

class UpdateCurrency(BaseModel):
    username: str
    currency: str

# Utility functions
def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

def generate_id() -> str:
    return str(uuid4())

def get_user_id_from_username(username: str) -> Optional[str]:
    """Get userId from username"""
    username_ref = db.collection('usernames').document(username).get()
    if username_ref.exists:
        return username_ref.to_dict().get('userId')
    return None

def get_user_by_id(user_id: str) -> Optional[dict]:
    """Get user document by userId"""
    user_ref = db.collection('users').document(user_id).get()
    if user_ref.exists:
        return user_ref.to_dict()
    return None

# ========== USER ENDPOINTS ==========

@app.post("/api/register")
async def register(user: UserRegister):
    """Register a new user with UUID"""
    
    # Check if username already exists
    existing_username = db.collection('usernames').document(user.username).get()
    if existing_username.exists:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Check if email already registered
    users = db.collection('users').where('email', '==', user.email).limit(1).stream()
    if len(list(users)) > 0:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Generate UUID for user
    user_id = generate_id()
    
    # Create user document with UUID
    db.collection('users').document(user_id).set({
        'userId': user_id,
        'username': user.username,
        'email': user.email,
        'password': hash_password(user.password),
        'fullName': user.fullName,
        'currency': user.currency,
        'savingsVault': 0,
        'createdAt': firestore.SERVER_TIMESTAMP
    })
    
    # Store username -> userId mapping
    db.collection('usernames').document(user.username).set({
        'userId': user_id,
        'username': user.username
    })
    
    return {"message": "User registered successfully", "username": user.username, "userId": user_id}

@app.post("/api/login")
async def login(credentials: UserLogin):
    """User login - returns userId"""
    user_id = get_user_id_from_username(credentials.username)
    
    if not user_id:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    user_data = get_user_by_id(user_id)
    
    if not user_data or user_data['password'] != hash_password(credentials.password):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    return {
        "message": "Login successful",
        "userId": user_id,
        "username": user_data['username'],
        "email": user_data['email'],
        "fullName": user_data['fullName'],
        "currency": user_data.get('currency', 'PKR')
    }

@app.get("/api/profile/{username}")
async def get_profile(username: str):
    """Get user profile"""
    user_id = get_user_id_from_username(username)
    
    if not user_id:
        raise HTTPException(status_code=404, detail="User not found")
    
    user_data = get_user_by_id(user_id)
    
    return {
        "userId": user_id,
        "username": user_data['username'],
        "email": user_data['email'],
        "fullName": user_data['fullName'],
        "currency": user_data.get('currency', 'PKR'),
        "savingsVault": user_data.get('savingsVault', 0)
    }

@app.put("/api/profile/name")
async def update_name(data: UpdateName):
    """Update user full name"""
    user_id = get_user_id_from_username(data.username)
    
    if not user_id:
        raise HTTPException(status_code=404, detail="User not found")
    
    db.collection('users').document(user_id).update({'fullName': data.fullName})
    return {"message": "Name updated successfully"}

@app.put("/api/profile/password")
async def update_password(data: UpdatePassword):
    """Update user password"""
    user_id = get_user_id_from_username(data.username)
    
    if not user_id:
        raise HTTPException(status_code=404, detail="User not found")
    
    user_data = get_user_by_id(user_id)
    
    if user_data['password'] != hash_password(data.oldPassword):
        raise HTTPException(status_code=400, detail="Old password is incorrect")
    
    db.collection('users').document(user_id).update({'password': hash_password(data.newPassword)})
    return {"message": "Password updated successfully"}

@app.put("/api/profile/currency")
async def update_currency(data: UpdateCurrency):
    """Update user currency preference"""
    user_id = get_user_id_from_username(data.username)
    
    if not user_id:
        raise HTTPException(status_code=404, detail="User not found")
    
    db.collection('users').document(user_id).update({'currency': data.currency})
    return {"message": "Currency updated successfully"}

# ========== TRANSACTION ENDPOINTS ==========

@app.post("/api/transactions")
async def add_transaction(txn: TransactionCreate):
    """Add a new transaction"""
    user_id = get_user_id_from_username(txn.username)
    
    if not user_id:
        raise HTTPException(status_code=404, detail="User not found")
    
    txn_id = generate_id()
    db.collection('transactions').document(txn_id).set({
        'transactionId': txn_id,
        'userId': user_id,
        'username': txn.username,  # Keep for display
        'type': txn.type,
        'category': txn.category,
        'amount': txn.amount,
        'description': txn.description,
        'date': txn.date,
        'currency': txn.currency,
        'created_at': datetime.now().isoformat()
    })
    return {"message": "Transaction added successfully", "id": txn_id}

@app.get("/api/transactions/{username}")
async def get_transactions(username: str, month: Optional[str] = None, limit: Optional[int] = None, all: Optional[bool] = False, group_by_month: Optional[bool] = False):
    """Get transactions for a user - supports monthly filtering and grouping"""
    user_id = get_user_id_from_username(username)
    
    if not user_id:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Get all transactions
    transactions = db.collection('transactions').where('userId', '==', user_id).stream()
    result = []
    for txn in transactions:
        data = txn.to_dict()
        
        # Filter by month only if specified and all=False
        if not all and month:
            if data['date'].startswith(month):
                data['id'] = data.get('transactionId', txn.id)
                result.append(data)
        else:
            # Return all transactions
            data['id'] = data.get('transactionId', txn.id)
            result.append(data)
    
    # Sort by date descending (newest first)
    result.sort(key=lambda x: x['date'], reverse=True)
    
    # Apply limit if specified
    if limit and limit > 0:
        result = result[:limit]
    
    # Group by month if requested
    if group_by_month:
        from collections import defaultdict
        grouped = defaultdict(list)
        for txn in result:
            month_key = txn['date'][:7]  # YYYY-MM
            grouped[month_key].append(txn)
        
        # Convert to list of monthly groups sorted by month descending
        monthly_data = []
        for month_key in sorted(grouped.keys(), reverse=True):
            monthly_data.append({
                'month': month_key,
                'transactions': grouped[month_key],
                'count': len(grouped[month_key]),
                'total_income': sum(t['amount'] for t in grouped[month_key] if t['type'] == 'income'),
                'total_expense': sum(t['amount'] for t in grouped[month_key] if t['type'] == 'expense')
            })
        
        return {"grouped_by_month": monthly_data}
    
    return result

@app.delete("/api/transactions/{txn_id}")
async def delete_transaction(txn_id: str):
    """Delete a transaction"""
    db.collection('transactions').document(txn_id).delete()
    return {"message": "Transaction deleted successfully"}

@app.get("/api/transactions/{username}/months")
async def get_available_months(username: str):
    """Get list of months that have transactions"""
    user_id = get_user_id_from_username(username)
    
    if not user_id:
        raise HTTPException(status_code=404, detail="User not found")
    
    transactions = db.collection('transactions').where('userId', '==', user_id).stream()
    months = set()
    for txn in transactions:
        data = txn.to_dict()
        month = data['date'][:7]  # YYYY-MM
        months.add(month)
    
    # Sort months descending (newest first)
    sorted_months = sorted(list(months), reverse=True)
    
    return {"months": sorted_months}

# ========== REPORT ENDPOINTS ==========

@app.get("/api/report/{username}")
async def get_monthly_report(username: str, month: Optional[str] = None):
    """Get monthly financial report - if no month provided, returns all-time data"""
    user_id = get_user_id_from_username(username)
    
    if not user_id:
        raise HTTPException(status_code=404, detail="User not found")
    
    transactions = db.collection('transactions').where('userId', '==', user_id).stream()
    txn_list = [txn.to_dict() for txn in transactions]
    
    # Filter by month if provided
    if month:
        filtered_txns = [t for t in txn_list if t['date'].startswith(month)]
    else:
        # All time data
        filtered_txns = txn_list
    
    total_income = sum(t['amount'] for t in filtered_txns if t['type'] == 'income')
    total_expense = sum(t['amount'] for t in filtered_txns if t['type'] == 'expense')
    
    category_breakdown = {}
    for txn in filtered_txns:
        if txn['type'] == 'expense':
            category_breakdown[txn['category']] = category_breakdown.get(txn['category'], 0) + txn['amount']
    
    user_data = get_user_by_id(user_id)
    
    # Get all transactions for overall balance
    all_txn_list = txn_list
    overall_income = sum(t['amount'] for t in all_txn_list if t['type'] == 'income')
    overall_expense = sum(t['amount'] for t in all_txn_list if t['type'] == 'expense')
    
    return {
        "month": month or "all-time",
        "total_income": total_income,
        "total_expense": total_expense,
        "balance": overall_income - overall_expense,  # Overall balance
        "monthly_balance": total_income - total_expense,  # Period balance
        "category_breakdown": category_breakdown,
        "transaction_count": len(filtered_txns),
        "savingsVault": user_data.get('savingsVault', 0)
    }

# ========== SAVINGS VAULT ENDPOINTS ==========

@app.post("/api/savings/add")
async def add_to_savings(data: SavingsOperation):
    """Add money to savings vault"""
    user_id = get_user_id_from_username(data.username)
    
    if not user_id:
        raise HTTPException(status_code=404, detail="User not found")
    
    user_data = get_user_by_id(user_id)
    current_savings = user_data.get('savingsVault', 0)
    new_balance = current_savings + data.amount
    
    db.collection('users').document(user_id).update({'savingsVault': new_balance})
    
    return {"message": "Added to savings", "newBalance": new_balance}

@app.post("/api/savings/withdraw")
async def withdraw_from_savings(data: SavingsOperation):
    """Withdraw from savings vault"""
    user_id = get_user_id_from_username(data.username)
    
    if not user_id:
        raise HTTPException(status_code=404, detail="User not found")
    
    user_data = get_user_by_id(user_id)
    current_savings = user_data.get('savingsVault', 0)
    
    if data.amount > current_savings:
        raise HTTPException(status_code=400, detail="Insufficient savings")
    
    new_balance = current_savings - data.amount
    db.collection('users').document(user_id).update({'savingsVault': new_balance})
    
    return {"message": "Withdrawn from savings", "newBalance": new_balance}

# ========== BUDGET ENDPOINTS ==========

@app.post("/api/budgets/set")
async def set_budget(budget: BudgetCreate):
    """Set a budget for a category"""
    user_id = get_user_id_from_username(budget.username)
    
    if not user_id:
        raise HTTPException(status_code=404, detail="User not found")
    
    budget_id = f"{user_id}_{budget.category}_{budget.month}"
    db.collection('budgets').document(budget_id).set({
        'budgetId': budget_id,
        'userId': user_id,
        'username': budget.username,
        'category': budget.category,
        'amount': budget.amount,
        'month': budget.month,
        'currency': budget.currency,
        'createdAt': firestore.SERVER_TIMESTAMP
    })
    return {"message": "Budget set successfully"}

@app.get("/api/budgets/status/{username}")
async def get_budget_status(username: str, month: Optional[str] = None):
    """Get budget status with spending"""
    user_id = get_user_id_from_username(username)
    
    if not user_id:
        raise HTTPException(status_code=404, detail="User not found")
    
    if not month:
        month = datetime.now().strftime("%Y-%m")
    
    budgets = db.collection('budgets').where('userId', '==', user_id).where('month', '==', month).stream()
    budget_list = [b.to_dict() for b in budgets]
    
    transactions = db.collection('transactions').where('userId', '==', user_id).stream()
    txn_list = [t.to_dict() for t in transactions if t.to_dict()['date'].startswith(month) and t.to_dict()['type'] == 'expense']
    
    budget_status = []
    for budget in budget_list:
        spent = sum(t['amount'] for t in txn_list if t['category'] == budget['category'])
        remaining = budget['amount'] - spent
        percentage = (spent / budget['amount'] * 100) if budget['amount'] > 0 else 0
        
        status = 'good'
        if percentage >= 100:
            status = 'exceeded'
        elif percentage >= 80:
            status = 'warning'
        
        budget_status.append({
            'budgetId': budget['budgetId'],  # Include budgetId for deletion
            'category': budget['category'],
            'budget': budget['amount'],
            'spent': spent,
            'remaining': remaining,
            'percentage': percentage,
            'status': status,
            'currency': budget['currency']
        })
    
    return {"budget_status": budget_status}

@app.delete("/api/budgets/{budget_id}")
async def delete_budget(budget_id: str):
    """Delete a budget"""
    db.collection('budgets').document(budget_id).delete()
    return {"message": "Budget deleted successfully"}

# ========== GOALS ENDPOINTS ==========

@app.post("/api/goals/create")
async def create_goal(goal: GoalCreate):
    """Create a financial goal"""
    user_id = get_user_id_from_username(goal.username)
    
    if not user_id:
        raise HTTPException(status_code=404, detail="User not found")
    
    goal_id = generate_id()
    
    db.collection('goals').document(goal_id).set({
        'goalId': goal_id,
        'userId': user_id,
        'username': goal.username,
        'name': goal.name,
        'target_amount': goal.target_amount,
        'current_amount': 0,
        'deadline': goal.deadline,
        'currency': goal.currency,
        'createdAt': firestore.SERVER_TIMESTAMP
    })
    
    return {"message": "Goal created successfully", "id": goal_id}

@app.get("/api/goals/{username}")
async def get_goals(username: str):
    """Get all financial goals for a user"""
    user_id = get_user_id_from_username(username)
    
    if not user_id:
        raise HTTPException(status_code=404, detail="User not found")
    
    goals = db.collection('goals').where('userId', '==', user_id).stream()
    goal_list = []
    
    for goal_doc in goals:
        goal = goal_doc.to_dict()
        
        progress = (goal['current_amount'] / goal['target_amount'] * 100) if goal['target_amount'] > 0 else 0
        remaining = goal['target_amount'] - goal['current_amount']
        
        deadline_date = datetime.strptime(goal['deadline'], "%Y-%m-%d")
        days_remaining = max(0, (deadline_date - datetime.now()).days)
        
        daily_required = remaining / days_remaining if days_remaining > 0 else 0
        
        goal_list.append({
            'id': goal.get('goalId', goal_doc.id),
            'name': goal['name'],
            'target_amount': goal['target_amount'],
            'current_amount': goal['current_amount'],
            'progress_percentage': progress,
            'remaining': remaining,
            'deadline': goal['deadline'],
            'days_remaining': days_remaining,
            'daily_savings_required': daily_required,
            'category': 'savings',
            'currency': goal['currency']
        })
    
    return goal_list

@app.post("/api/goals/contribute")
async def contribute_to_goal(data: ContributeGoal):
    """Contribute money to a goal - deducts from balance via expense transaction"""
    goal_ref = db.collection('goals').document(data.id)
    goal_doc = goal_ref.get()
    
    if not goal_doc.exists:
        raise HTTPException(status_code=404, detail="Goal not found")
    
    goal_data = goal_doc.to_dict()
    user_id = goal_data['userId']
    username = goal_data['username']
    
    # Get user's current balance (calculated from transactions)
    transactions = db.collection('transactions').where('userId', '==', user_id).stream()
    txn_list = [txn.to_dict() for txn in transactions]
    
    total_income = sum(t['amount'] for t in txn_list if t['type'] == 'income')
    total_expense = sum(t['amount'] for t in txn_list if t['type'] == 'expense')
    current_balance = total_income - total_expense
    
    # Check if user has sufficient balance
    if current_balance < data.amount:
        raise HTTPException(status_code=400, detail=f"Insufficient balance. Available: {current_balance:.2f}")
    
    # Create expense transaction to deduct from balance
    txn_id = generate_id()
    db.collection('transactions').document(txn_id).set({
        'transactionId': txn_id,
        'userId': user_id,
        'username': username,
        'type': 'expense',
        'category': 'Savings',
        'amount': data.amount,
        'description': f"Contribution to goal: {goal_data['name']}",
        'date': datetime.now().strftime("%Y-%m-%d"),
        'currency': goal_data.get('currency', 'PKR'),
        'created_at': datetime.now().isoformat()
    })
    
    # Update goal amount
    new_amount = goal_data['current_amount'] + data.amount
    goal_ref.update({'current_amount': new_amount})
    
    # New balance after expense
    new_balance = current_balance - data.amount
    
    return {
        "message": "Contribution added successfully",
        "new_goal_amount": new_amount,
        "new_balance": new_balance
    }

@app.delete("/api/goals/{goal_id}")
async def delete_goal(goal_id: str, completed: bool = False):
    """Delete a financial goal - returns money to balance if cancelled (not completed)"""
    goal_ref = db.collection('goals').document(goal_id)
    goal_doc = goal_ref.get()
    
    if not goal_doc.exists:
        raise HTTPException(status_code=404, detail="Goal not found")
    
    goal_data = goal_doc.to_dict()
    user_id = goal_data['userId']
    username = goal_data['username']
    current_amount = goal_data['current_amount']
    target_amount = goal_data['target_amount']
    goal_name = goal_data['name']
    
    # Check if goal is actually complete
    is_actually_complete = current_amount >= target_amount
    
    message = ""
    returned_amount = 0
    
    # If user says it's NOT completed and there's money contributed, return it
    if not completed and current_amount > 0:
        # Create income transaction to return money to balance
        txn_id = generate_id()
        db.collection('transactions').document(txn_id).set({
            'transactionId': txn_id,
            'userId': user_id,
            'username': username,
            'type': 'income',
            'category': 'Goal Refund',
            'amount': current_amount,
            'description': f"Refund from cancelled goal: {goal_name}",
            'date': datetime.now().strftime("%Y-%m-%d"),
            'currency': goal_data.get('currency', 'PKR'),
            'created_at': datetime.now().isoformat()
        })
        
        returned_amount = current_amount
        message = f"Goal cancelled. {current_amount:.2f} returned to balance."
    elif completed:
        if is_actually_complete:
            message = "🎉 Congratulations! Goal completed and removed!"
        else:
            message = f"Goal marked as completed and removed. (Reached {current_amount:.2f} of {target_amount:.2f})"
    else:
        message = "Goal deleted."
    
    # Delete the goal
    goal_ref.delete()
    
    return {
        "message": message,
        "was_complete": is_actually_complete,
        "returned_amount": returned_amount
    }

# ========== CURRENCY ENDPOINTS ==========

@app.get("/api/currency/rates")
async def get_currency_rates():
    """Get currency exchange rates"""
    rates_ref = db.collection('currency_rates').document('rates')
    rates_doc = rates_ref.get()
    
    if not rates_doc.exists:
        default_rates = {
            "USD": 1.0,
            "EUR": 0.85,
            "GBP": 0.73,
            "PKR": 278.50,
            "SAR": 3.75,
            "AED": 3.67
        }
        rates_ref.set(default_rates)
        return {"rates": default_rates}
    
    return {"rates": rates_doc.to_dict()}

# ========== EXPORT ENDPOINTS ==========

@app.get("/api/export/{username}")
async def export_to_excel(username: str, month: Optional[str] = None):
    """Export transactions to Excel"""
    user_id = get_user_id_from_username(username)
    
    if not user_id:
        raise HTTPException(status_code=404, detail="User not found")
    
    transactions = db.collection('transactions').where('userId', '==', user_id).stream()
    txn_list = [txn.to_dict() for txn in transactions]
    
    if month:
        txn_list = [t for t in txn_list if t['date'].startswith(month)]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    
    headers = ["Date", "Type", "Category", "Amount", "Currency", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row, txn in enumerate(txn_list, 2):
        ws.cell(row=row, column=1, value=txn['date'])
        ws.cell(row=row, column=2, value=txn['type'])
        ws.cell(row=row, column=3, value=txn['category'])
        ws.cell(row=row, column=4, value=txn['amount'])
        ws.cell(row=row, column=5, value=txn.get('currency', 'PKR'))
        ws.cell(row=row, column=6, value=txn['description'])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"transactions_{username}_{month or 'all'}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ========== AI CHAT ENDPOINTS ==========

@app.post("/api/chat")
async def chat_with_ai(message: ChatMessage):
    """Chat with AI assistant"""
    if not GOOGLE_API_KEY:
        raise HTTPException(status_code=500, detail="AI service not configured")
    
    user_id = get_user_id_from_username(message.username)
    
    if not user_id:
        raise HTTPException(status_code=404, detail="User not found")
    
    transactions = db.collection('transactions').where('userId', '==', user_id).stream()
    txn_list = [txn.to_dict() for txn in transactions]
    
    total_income = sum(t['amount'] for t in txn_list if t['type'] == 'income')
    total_expense = sum(t['amount'] for t in txn_list if t['type'] == 'expense')
    
    context = f"""
    User's Financial Summary:
    - Total Income: ${total_income}
    - Total Expenses: ${total_expense}
    - Balance: ${total_income - total_expense}
    - Number of Transactions: {len(txn_list)}
    
    User Question: {message.message}
    
    Provide helpful financial advice based on this data.
    """
    
    response = client.models.generate_content(
        model='gemini-2.5-flash',
        contents=context
    )
    
    return {
        "response": response.text,
        "context": {
            "income": total_income,
            "expense": total_expense,
            "balance": total_income - total_expense
        }
    }

@app.get("/")
async def root():
    return {
        "message": "Expense Tracker API with Firestore (UUID-based)",
        "version": "3.0.0",
        "status": "running",
        "database": "Firestore",
        "schema": "UUID-based user tracking"
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
